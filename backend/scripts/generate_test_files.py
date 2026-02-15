# Copyright (c) 2026 Aaron Guo. All rights reserved.
# Use of this source code is governed by the proprietary license
# found in the LICENSE file in the root directory of this source tree.

"""
Generate complex Excel test files to mimic chaotic factory environments.
Tests the LineSight parser's ability to handle real-world messy data.
"""

import random
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# Create output directory
output_dir = Path("../sample_data/complex")
output_dir.mkdir(exist_ok=True, parents=True)

print("🏭 Generating Complex Factory Excel Files...")
print("=" * 60)

# ============================================================================
# 1. PRODUCTION TRACKING (Multi-Header Chaos)
# ============================================================================
print("\n1️⃣  Creating: production_tracking_multiheader.xlsx")

wb = Workbook()
ws = wb.active
ws.title = "Production Summary"

# Title rows (merged cells)
ws.merge_cells("A1:H1")
ws["A1"] = "PRODUCTION SUMMARY - WEEK 51 (Dec 2024)"
ws["A1"].font = Font(bold=True, size=14)
ws["A1"].alignment = Alignment(horizontal="center")

ws.merge_cells("A2:H2")
ws["A2"] = "Factory: Dhaka Garments Ltd. | Line: A-01"
ws["A2"].font = Font(italic=True, size=10)
ws["A2"].alignment = Alignment(horizontal="center")

# Headers in row 3
headers = [
    "Date",
    "Style#",
    "PO Number",
    "Target Qty",
    "Actual Output",
    "Operators",
    "Defects",
    "DHU%",
]
for col, header in enumerate(headers, start=1):
    cell = ws.cell(row=3, column=col, value=header)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

# Data rows with mixed date formats
date_formats = [
    lambda d: d.strftime("%m/%d/%Y"),  # 12/25/2024
    lambda d: d.strftime("%d-%b-%y"),  # 25-Dec-24
    lambda d: d.strftime("%Y-%m-%d"),  # 2024-12-25
]

start_date = datetime(2024, 12, 23)
styles = ["ABC-123", "XYZ-789", "DEF-456"]
pos = ["PO-2024-001", "PO-2024-002", "PO-2024-003"]

for day in range(5):
    current_date = start_date + timedelta(days=day)
    date_str = random.choice(date_formats)(current_date)

    style = random.choice(styles)
    po = random.choice(pos)
    target = random.randint(400, 600)
    actual = int(target * random.uniform(0.85, 1.05))
    operators = random.randint(25, 35)
    defects = (
        random.randint(5, 25) if random.random() > 0.3 else ""
    )  # Sometimes missing
    dhu = round((defects / actual * 100), 2) if defects else ""

    row = [date_str, style, po, target, actual, operators, defects, dhu]
    ws.append(row)

wb.save(output_dir / "production_tracking_multiheader.xlsx")
print("   ✅ Created with headers in row 3, merged title cells, mixed date formats")

# ============================================================================
# 2. QUALITY INSPECTION (Multi-Sheet)
# ============================================================================
print("\n2️⃣  Creating: quality_inspection_multisheet.xlsx")

wb = Workbook()

# Sheet 1: In-Line QC
ws1 = wb.active
ws1.title = "In-Line QC"
inline_data = {
    "Date": ["12/23/2024", "12/23/2024", "12/24/2024", "12/24/2024", "12/25/2024"],
    "Line": ["A-01", "A-02", "A-01", "A-02", "A-01"],
    "Checked": [120, 150, 135, 140, 125],
    "Defects": [8, 12, 7, 15, 9],
    "DHU": [6.67, 8.0, 5.19, 10.71, 7.2],  # Sometimes calculated
}
df1 = pd.DataFrame(inline_data)
for r in dataframe_to_rows(df1, index=False, header=True):
    ws1.append(r)

# Sheet 2: End-Line QC
ws2 = wb.create_sheet("End-Line QC")
endline_data = {
    "PO": ["PO-2024-001", "PO-2024-002", "PO-2024-003"],
    "Style": ["ABC-123", "XYZ-789", "DEF-456"],
    "Color": ["Navy", "Black", "White"],
    "Inspected": [500, 450, 600],
    "Pass": [475, 430, 570],
    "Fail": [25, 20, 30],
    "Pass Rate": ["95%", "95.6%", "0.95"],  # Mixed formats
}
df2 = pd.DataFrame(endline_data)
for r in dataframe_to_rows(df2, index=False, header=True):
    ws2.append(r)

# Sheet 3: Summary
ws3 = wb.create_sheet("Summary")
summary_data = {
    "Week": ["Week 50", "Week 51"],
    "Total Checked": [2500, 2700],
    "Total Defects": [180, 195],
    "Avg DHU": [7.2, 7.22],
}
df3 = pd.DataFrame(summary_data)
for r in dataframe_to_rows(df3, index=False, header=True):
    ws3.append(r)

wb.save(output_dir / "quality_inspection_multisheet.xlsx")
print("   ✅ Created with 3 sheets: In-Line QC, End-Line QC, Summary")

# ============================================================================
# 3. WORKFORCE ATTENDANCE (Pivot Format)
# ============================================================================
print("\n3️⃣  Creating: workforce_attendance_pivot.xlsx")

wb = Workbook()
ws = wb.active
ws.title = "Attendance"

# Headers
ws["A1"] = "Operator Name"
dates = [
    (datetime(2024, 12, 23) + timedelta(days=i)).strftime("%m/%d") for i in range(7)
]
for col, date in enumerate(dates, start=2):
    ws.cell(row=1, column=col, value=date)

# Operators
operators = [
    "John Doe",
    "Jane Smith",
    "রহিম আলী",
    "Fatima Khan",  # Mixed languages
    "Michael Chen",
    "Sarah Johnson",
    "আব্দুল করিম",
]

attendance_codes = ["P", "A", "L", "H"]  # Present, Absent, Leave, Holiday

for row, operator in enumerate(operators, start=2):
    ws.cell(row=row, column=1, value=operator)
    for col in range(2, 9):
        code = random.choice(attendance_codes)
        ws.cell(row=row, column=col, value=code)

wb.save(output_dir / "workforce_attendance_pivot.xlsx")
print("   ✅ Created with pivot format (operators in rows, dates in columns)")

# ============================================================================
# 4. FABRIC TRACEABILITY (UFLPA Compliance)
# ============================================================================
print("\n4️⃣  Creating: fabric_traceability_uflpa.xlsx")

wb = Workbook()
ws = wb.active
ws.title = "Fabric Traceability"

headers = [
    "Lot Number",
    "Fabric Type",
    "Composition",
    "Origin Country",
    "Supplier",
    "Mill Country",
    "Cert Number",
    "Received Date",
]
ws.append(headers)

traceability_data = [
    [
        "F-2024-001",
        "Cotton Jersey",
        "100% Cotton",
        "India",
        "ABC Textiles",
        "India",
        "GOTS-12345",
        "11/15/2024",
    ],
    [
        "F-2024-001-A",
        "Cotton Jersey",
        "100% Cotton",
        "India → Vietnam",
        "ABC → XYZ Mills",
        "India",
        "GOTS-12345",
        "11/20/2024",
    ],
    [
        "F-2024-002",
        "Polyester Blend",
        "65% Poly 35% Cotton",
        "China",
        "DEF Fabrics",
        "China",
        "OEKO-67890",
        "11/18/2024",
    ],
    [
        "F-2024-003",
        "Organic Cotton",
        "100% Organic Cotton",
        "Turkey",
        "GHI Textiles",
        "Turkey",
        "GOTS-11111",
        "11/22/2024",
    ],
    [
        "F-2024-004",
        "Recycled Poly",
        "100% rPET",
        "Vietnam",
        "JKL Recycling",
        "Vietnam",
        "GRS-22222",
        "11/25/2024",
    ],
]

for row in traceability_data:
    ws.append(row)

# Highlight nested lot numbers
ws["A3"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

wb.save(output_dir / "fabric_traceability_uflpa.xlsx")
print("   ✅ Created with nested lot numbers and multi-country origins")

# ============================================================================
# 5. MIXED FORMAT CHAOS
# ============================================================================
print("\n5️⃣  Creating: mixed_format_chaos.xlsx")

wb = Workbook()
ws = wb.active
ws.title = "Production Data"

# Random empty rows
ws.append([])
ws.append(["Production Report - CONFIDENTIAL"])
ws.append([])

# Headers with comments
headers = ["Date", "Style", "Qty", "Status", "Notes"]
ws.append(headers)

# Data with chaos
data_rows = [
    ["12/23/2024", "ABC-123", 500, "Complete", ""],
    ["", "", "", "", ""],  # Empty row
    [
        "12/24/2024",
        "XYZ-789",
        "=500+50",
        "In Progress",
        "CHECK THIS!",
    ],  # Formula + comment
    ["25-Dec-24", "DEF-456", "৬০০", "Complete", ""],  # Bengali number
    ["2024-12-26", "GHI-012", 450, "Delayed", "Fabric shortage"],
    ["", "", "", "", ""],  # Empty row
    ["12/27/2024", "JKL-345", "550 pcs", "Complete", ""],  # Unit in number
]

for row in data_rows:
    ws.append(row)

# Color-code status
ws["D5"].fill = PatternFill(
    start_color="00FF00", end_color="00FF00", fill_type="solid"
)  # Green
ws["D6"].fill = PatternFill(
    start_color="FFFF00", end_color="FFFF00", fill_type="solid"
)  # Yellow
ws["D8"].fill = PatternFill(
    start_color="FF0000", end_color="FF0000", fill_type="solid"
)  # Red

wb.save(output_dir / "mixed_format_chaos.xlsx")
print("   ✅ Created with empty rows, formulas, comments, color coding, Bengali text")

# ============================================================================
# 6. MINIMAL DATA (Edge Case)
# ============================================================================
print("\n6️⃣  Creating: minimal_data_edge_case.xlsx")

minimal_data = {
    "Style": ["ABC-123", "XYZ-789", "DEF-456"],
    "Qty": [500, 300, 450],
    "Color": ["Navy", "Black", "White"],
}

df = pd.DataFrame(minimal_data)
df.to_excel(output_dir / "minimal_data_edge_case.xlsx", index=False)
print("   ✅ Created with only 3 columns (tests graceful degradation)")

print("\n" + "=" * 60)
print("✅ All 6 complex Excel files generated successfully!")
print(f"📁 Location: {output_dir.absolute()}")
print("\nFiles created:")
print("  1. production_tracking_multiheader.xlsx")
print("  2. quality_inspection_multisheet.xlsx")
print("  3. workforce_attendance_pivot.xlsx")
print("  4. fabric_traceability_uflpa.xlsx")
print("  5. mixed_format_chaos.xlsx")
print("  6. minimal_data_edge_case.xlsx")
