# Copyright (c) 2026 Aaron Guo. All rights reserved.
# Use of this source code is governed by the proprietary license
# found in the LICENSE file in the root directory of this source tree.

"""
Test the FlexibleExcelParser and SemanticETLAgent against complex test files.
Validates the AI architecture and decision logging.
"""

import sys
from pathlib import Path

sys.path.append(str(Path(__file__).parent))

from openpyxl import load_workbook

from app.services.excel_parser import FlexibleExcelParser
from app.services.llm_agent import SemanticETLAgent


def test_file(file_path: Path, use_ai: bool = False):
    """Test a single Excel file with the parser."""
    print(f"\n{'=' * 80}")
    print(f"📄 Testing: {file_path.name}")
    print(f"🤖 AI Mode: {'ENABLED' if use_ai else 'DISABLED (Fuzzy Matching Only)'}")
    print(f"{'=' * 80}")

    try:
        parser = FlexibleExcelParser(strict_mode=False)
        result = parser.parse_file(file_path)

        print(f"\n✅ Parse Success: {result.success}")
        print(f"📊 Records Created: {len(result.records)}")
        print(f"🗺️  Column Mappings: {len(result.column_mappings)}")
        print(f"⚠️  Warnings: {len(result.warnings)}")
        print(f"❌ Errors: {len(result.errors)}")

        # Show column mappings
        print("\n🔗 Column Mappings:")
        for mapping in result.column_mappings:
            print(
                f"   {mapping.source_column:20} → {mapping.target_field:20} ({mapping.confidence.value})"
            )

        # Show decision logs
        if result.decision_logs:
            print("\n📝 Decision Logs:")
            for log in result.decision_logs[:5]:  # Show first 5
                print(f"   • {log}")
            if len(result.decision_logs) > 5:
                print(f"   ... and {len(result.decision_logs) - 5} more")

        # Show sample records
        if result.records:
            print("\n📋 Sample Records (first 2):")
            for i, record in enumerate(result.records[:2], 1):
                print(f"\n   Record {i}:")
                for key, value in record.items():
                    if value is not None:
                        print(f"      {key}: {value}")

        # Show warnings and errors
        if result.warnings:
            print("\n⚠️  Warnings:")
            for warning in result.warnings:
                print(f"   • {warning}")

        if result.errors:
            print("\n❌ Errors:")
            for error in result.errors:
                print(f"   • {error}")

        return result

    except Exception as e:
        print(f"\n❌ EXCEPTION: {type(e).__name__}: {str(e)}")
        import traceback

        traceback.print_exc()
        return None


def test_ai_schema_inference(file_path: Path):
    """Test AI schema inference on a file."""
    print(f"\n{'=' * 80}")
    print(f"🤖 AI SCHEMA INFERENCE: {file_path.name}")
    print(f"{'=' * 80}")

    try:
        # Load sample rows
        wb = load_workbook(file_path)
        ws = wb.active

        sample_rows = []
        for row in ws.iter_rows(max_row=20, values_only=True):
            sample_rows.append(list(row))

        # Run AI inference
        agent = SemanticETLAgent()
        schema = agent.infer_schema(
            sample_rows=sample_rows,
            filename=file_path.name,
            file_type_hint="production_data",
        )

        print("\n✅ Schema Inference Complete")
        print(f"📍 Header Row: {schema.header_row}")
        print(f"📋 Detected Headers: {schema.detected_headers}")

        print("\n🗺️  Column Mappings:")
        for source, target in schema.column_mappings.items():
            confidence = schema.confidence_scores.get(source, 0.0)
            print(f"   {source:20} → {target:20} (confidence: {confidence:.2f})")

        print("\n💡 Recommendations:")
        for rec in schema.recommendations:
            print(f"   • {rec}")

        if schema.suggested_widgets:
            print("\n🎨 Suggested Widgets:")
            for widget in schema.suggested_widgets:
                print(f"   • {widget}")

        return schema

    except Exception as e:
        print(f"\n❌ AI EXCEPTION: {type(e).__name__}: {str(e)}")
        import traceback

        traceback.print_exc()
        return None


def main():
    """Run tests on all complex Excel files."""
    test_dir = Path(__file__).parent / "../sample_data/complex"

    if not test_dir.exists():
        print(f"❌ Test directory not found: {test_dir}")
        return

    test_files = list(test_dir.glob("*.xlsx"))

    if not test_files:
        print(f"❌ No test files found in {test_dir}")
        return

    print("\n🏭 LineSight Parser Test Suite")
    print(f"📁 Test Directory: {test_dir.absolute()}")
    print(f"📊 Test Files: {len(test_files)}")

    results = {}

    # Test each file with fuzzy matching (no AI)
    print(f"\n\n{'#' * 80}")
    print("# PHASE 1: FUZZY MATCHING (NO AI)")
    print(f"{'#' * 80}")

    for file_path in sorted(test_files):
        result = test_file(file_path, use_ai=False)
        results[file_path.name] = {"fuzzy": result}

    # Test AI schema inference on select files
    print(f"\n\n{'#' * 80}")
    print("# PHASE 2: AI SCHEMA INFERENCE")
    print(f"{'#' * 80}")

    ai_test_files = [
        "production_tracking_multiheader.xlsx",
        "quality_inspection_multisheet.xlsx",
        "mixed_format_chaos.xlsx",
    ]

    for filename in ai_test_files:
        file_path = test_dir / filename
        if file_path.exists():
            schema = test_ai_schema_inference(file_path)
            if filename in results:
                results[filename]["ai_schema"] = schema

    # Summary
    print(f"\n\n{'#' * 80}")
    print("# TEST SUMMARY")
    print(f"{'#' * 80}")

    print(f"\n{'File':<40} {'Fuzzy Parse':<15} {'Records':<10}")
    print(f"{'-' * 80}")

    for filename, result_dict in results.items():
        fuzzy_result = result_dict.get("fuzzy")
        status = "✅ SUCCESS" if fuzzy_result and fuzzy_result.success else "❌ FAILED"
        record_count = len(fuzzy_result.records) if fuzzy_result else 0
        print(f"{filename:<40} {status:<15} {record_count:<10}")

    print("\n✅ Test suite complete!")


if __name__ == "__main__":
    main()
